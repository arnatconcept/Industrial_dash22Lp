import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def generar_plantilla_excel(model_name, fields, choices, file_path):
    """
    Genera un archivo Excel con la estructura de un modelo específico
    
    Args:
        model_name (str): Nombre del modelo
        fields (list): Lista de diccionarios con información de los campos
        choices (dict): Diccionario con opciones para campos de selección
        file_path (str): Ruta donde guardar el archivo
    """
    # Crear libro de Excel
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Datos"
    
    # Hoja de instrucciones
    ws_instructions = wb.create_sheet("Instrucciones")
    
    # Escribir encabezados en hoja de datos
    headers = [field['name'] for field in fields]
    ws_data.append(headers)
    
    # Configurar estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Aplicar estilos a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        column_letter = get_column_letter(col_num)
        ws_data.column_dimensions[column_letter].width = max(15, len(header) * 1.2)
    
    # Agregar validaciones de datos
    for col_num, field in enumerate(fields, 1):
        if field.get('choices'):
            # Crear validación para campos de selección
            dv = DataValidation(
                type="list", 
                formula1=f'"{",".join(field["choices"])}"', 
                allow_blank=field.get('blank', False)
            )
            ws_data.add_data_validation(dv)
            dv.add(f"{get_column_letter(col_num)}2:{get_column_letter(col_num)}1048576")
    
    # Escribir hoja de instrucciones
    ws_instructions.append(["INSTRUCCIONES PARA CARGA DE DATOS"])
    ws_instructions.append([])
    ws_instructions.append(["Modelo:", model_name])
    ws_instructions.append(["Fecha generación:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_instructions.append([])
    
    # Tabla de campos
    ws_instructions.append(["CAMPO", "TIPO", "REQUERIDO", "DESCRIPCIÓN", "EJEMPLO", "OPCIONES"])
    for field in fields:
        ws_instructions.append([
            field['name'],
            field['type'],
            "No" if field.get('blank') else "Sí",
            field.get('help_text', ''),
            field.get('example', ''),
            ", ".join(field['choices']) if field.get('choices') else ''
        ])
    
    # Estilo para hoja de instrucciones
    for row in ws_instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_instructions.column_dimensions[col].width = 20
    
    # Guardar archivo
    wb.save(file_path)

def generar_plantillas_completas(output_dir="plantillas_excel"):
    """Genera todas las plantillas necesarias para el sistema"""
    
    # Crear directorio si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Definir estructura para cada modelo
    modelos = [
        {
            'name': 'LineaProduccion',
            'fields': [
                {'name': 'nombre', 'type': 'CharField', 'help_text': 'Nombre único de la línea de producción', 'example': 'Línea 1', 'blank': False},
                {'name': 'descripcion', 'type': 'TextField', 'help_text': 'Descripción opcional', 'example': 'Línea principal de ensamblaje', 'blank': True},
            ]
        },
        {
            'name': 'Sector',
            'fields': [
                {'name': 'nombre', 'type': 'CharField', 'help_text': 'Nombre del sector', 'example': 'Soldadura', 'blank': False},
                {'name': 'linea', 'type': 'ForeignKey', 'help_text': 'Nombre de la línea de producción a la que pertenece', 'example': 'Línea 1', 'blank': False},
            ]
        },
        {
            'name': 'Equipo',
            'fields': [
                {'name': 'nombre', 'type': 'CharField', 'help_text': 'Nombre del equipo', 'example': 'Máquina de soldar', 'blank': False},
                {'name': 'sector', 'type': 'ForeignKey', 'help_text': 'Nombre del sector al que pertenece (formato: "Línea - Sector")', 'example': 'Línea 1 - Soldadura', 'blank': False},
            ]
        },
        {
            'name': 'Deposito',
            'fields': [
                {'name': 'nombre', 'type': 'CharField', 'help_text': 'Nombre único del depósito', 'example': 'Depósito Central', 'blank': False},
                {'name': 'ubicacion', 'type': 'CharField', 'help_text': 'Ubicación física del depósito', 'example': 'Edificio B, Planta Baja', 'blank': True},
            ]
        },
        {
            'name': 'Motor',
            'fields': [
                {'name': 'codigo', 'type': 'CharField', 'help_text': 'Código único identificador', 'example': 'MOT-001', 'blank': False},
                {'name': 'potencia', 'type': 'CharField', 'help_text': 'Potencia en kW o HP', 'example': '5.5 kW', 'blank': False},
                {'name': 'tipo', 'type': 'CharField', 'help_text': 'Tipo de motor', 'example': 'Trifásico', 'blank': False},
                {'name': 'rpm', 'type': 'CharField', 'help_text': 'Revoluciones por minuto', 'example': '1450 rpm', 'blank': False},
                {'name': 'brida', 'type': 'CharField', 'help_text': 'Tipo de brida', 'example': 'B5', 'blank': False},
                {'name': 'anclaje', 'type': 'CharField', 'help_text': 'Tipo de anclaje', 'example': 'Pie', 'blank': False},
                {'name': 'estado', 'type': 'CharField', 'choices': ['operativo', 'reparacion', 'baja', 'standby'], 'help_text': 'Estado actual del motor', 'blank': False},
                {'name': 'ubicacion_tipo', 'type': 'CharField', 'choices': ['linea', 'deposito', 'mantenimiento'], 'help_text': 'Tipo de ubicación', 'blank': False},
                {'name': 'linea', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea', 'example': 'Línea 1', 'blank': True},
                {'name': 'sector', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea (formato: "Línea - Sector")', 'example': 'Línea 1 - Soldadura', 'blank': True},
                {'name': 'equipo', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea (formato: "Línea - Sector - Equipo")', 'example': 'Línea 1 - Soldadura - Máquina 1', 'blank': True},
                {'name': 'deposito', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=deposito', 'example': 'Depósito Central', 'blank': True},
                {'name': 'fecha_instalacion', 'type': 'DateField', 'help_text': 'Fecha de instalación (YYYY-MM-DD)', 'example': '2023-01-15', 'blank': True},
                {'name': 'horas_uso', 'type': 'IntegerField', 'help_text': 'Horas de uso acumuladas', 'example': '1500', 'blank': True},
            ]
        },
        {
            'name': 'Variador',
            'fields': [
                {'name': 'codigo', 'type': 'CharField', 'help_text': 'Código único identificador', 'example': 'VAR-001', 'blank': False},
                {'name': 'marca', 'type': 'CharField', 'help_text': 'Marca del variador', 'example': 'Siemens', 'blank': False},
                {'name': 'modelo', 'type': 'CharField', 'help_text': 'Modelo específico', 'example': 'G120', 'blank': False},
                {'name': 'potencia', 'type': 'CharField', 'help_text': 'Potencia en kW', 'example': '7.5 kW', 'blank': False},
                {'name': 'estado', 'type': 'CharField', 'choices': ['operativo', 'reparacion', 'baja', 'standby'], 'help_text': 'Estado actual', 'blank': False},
                {'name': 'ubicacion_tipo', 'type': 'CharField', 'choices': ['linea', 'deposito', 'mantenimiento'], 'help_text': 'Tipo de ubicación', 'blank': False},
                {'name': 'linea', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea', 'example': 'Línea 1', 'blank': True},
                {'name': 'sector', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea (formato: "Línea - Sector")', 'example': 'Línea 1 - Soldadura', 'blank': True},
                {'name': 'equipo', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=linea (formato: "Línea - Sector - Equipo")', 'example': 'Línea 1 - Soldadura - Máquina 1', 'blank': True},
                {'name': 'deposito', 'type': 'ForeignKey', 'help_text': 'Requerido si ubicacion_tipo=deposito', 'example': 'Depósito Central', 'blank': True},
                {'name': 'fecha_instalacion', 'type': 'DateField', 'help_text': 'Fecha de instalación (YYYY-MM-DD)', 'example': '2023-01-15', 'blank': True},
                {'name': 'horas_uso', 'type': 'IntegerField', 'help_text': 'Horas de uso acumuladas', 'example': '1500', 'blank': True},
            ]
        },
        {
            'name': 'OrdenMantenimiento',
            'fields': [
                {'name': 'titulo', 'type': 'CharField', 'help_text': 'Título descriptivo', 'example': 'Revisión preventiva motor principal', 'blank': False},
                {'name': 'descripcion', 'type': 'TextField', 'help_text': 'Descripción detallada', 'example': 'Revisión trimestral según plan', 'blank': False},
                {'name': 'tipo', 'type': 'CharField', 'choices': ['preventivo', 'correctivo', 'predictivo', 'calibracion'], 'help_text': 'Tipo de mantenimiento', 'blank': False},
                {'name': 'prioridad', 'type': 'CharField', 'choices': ['baja', 'media', 'alta', 'critica'], 'help_text': 'Prioridad de la orden', 'blank': False},
                {'name': 'estado', 'type': 'CharField', 'choices': ['pendiente', 'asignada', 'en_proceso', 'completada', 'cancelada'], 'help_text': 'Estado actual', 'blank': False},
                {'name': 'operario_asignado', 'type': 'ForeignKey', 'help_text': 'Usuario asignado (username)', 'example': 'jperez', 'blank': True},
                {'name': 'fecha_creacion', 'type': 'DateTimeField', 'help_text': 'Fecha de creación (YYYY-MM-DD HH:MM)', 'example': '2023-05-20 09:30', 'blank': False},
                {'name': 'fecha_cierre', 'type': 'DateTimeField', 'help_text': 'Fecha de cierre (YYYY-MM-DD HH:MM)', 'example': '2023-05-21 16:45', 'blank': True},
                {'name': 'tiempo_estimado', 'type': 'IntegerField', 'help_text': 'Tiempo estimado en minutos', 'example': '120', 'blank': True},
                {'name': 'tiempo_real', 'type': 'IntegerField', 'help_text': 'Tiempo real en minutos', 'example': '135', 'blank': True},
                {'name': 'motores', 'type': 'ManyToManyField', 'help_text': 'Códigos de motores (separados por coma)', 'example': 'MOT-001, MOT-002', 'blank': True},
                {'name': 'variadores', 'type': 'ManyToManyField', 'help_text': 'Códigos de variadores (separados por coma)', 'example': 'VAR-001', 'blank': True},
            ]
        },
        {
            'name': 'Proveedor',
            'fields': [
                {'name': 'nombre', 'type': 'CharField', 'help_text': 'Nombre del proveedor', 'example': 'Técnica Industrial S.A.', 'blank': False},
                {'name': 'especialidad', 'type': 'CharField', 'choices': ['electrico', 'mecanico', 'electronico', 'hidraulico', 'neumatico', 'general'], 'help_text': 'Especialidad principal', 'blank': False},
                {'name': 'contacto', 'type': 'CharField', 'help_text': 'Persona de contacto', 'example': 'Juan Pérez', 'blank': False},
                {'name': 'telefono', 'type': 'CharField', 'help_text': 'Teléfono de contacto', 'example': '+54 11 1234-5678', 'blank': False},
                {'name': 'email', 'type': 'EmailField', 'help_text': 'Correo electrónico', 'example': 'contacto@proveedor.com', 'blank': False},
                {'name': 'direccion', 'type': 'TextField', 'help_text': 'Dirección física', 'example': 'Av. Siempreviva 123, CABA', 'blank': True},
            ]
        },
    ]
    
    # Generar plantillas para cada modelo
    for modelo in modelos:
        file_path = os.path.join(output_dir, f"plantilla_{modelo['name']}.xlsx")
        generar_plantilla_excel(
            model_name=modelo['name'],
            fields=modelo['fields'],
            choices={},
            file_path=file_path
        )
        print(f"Plantilla generada: {file_path}")
    
    print("\n¡Todas las plantillas han sido generadas con éxito!")

if __name__ == "__main__":
    generar_plantillas_completas()