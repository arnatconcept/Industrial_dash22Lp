import openpyxl
from openpyxl.styles import Font

# Lista de sectores únicos
sectores = [
    "PUPITRE",
    "APILADO",
    "QP1",
    "QP2",
    "PUENTE L2",
    "DRAGA",
    "MOLAZA",
    "PUENTE L1",
    "PUENTE L1/PUENTE L2",
    "MANUTENCIÓN",
    "DESAPILADO",
    "PREPARACIÓN",
    "CARGADOR",
    "FABRICACIÓN",
    "SECADERO",
    "HORNO",
    "Servicios Auxiliares",
    "SECADERO VENT",
    "Barredora",
    "SECADERO CTIBOR V4",
    "SECADERO OMEGA V6",
    "SECADERO OMEGA V8",
    "SECADERO CTIBOR V2",
    "Manut. Beralmar"
]

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sectores"

# Encabezados
ws['A1'] = "id"
ws['B1'] = "nombre"
ws['C1'] = "linea_id"

# Estilo para encabezados
for cell in ws[1]:
    cell.font = Font(bold=True)

# Llenar datos
for i, sector in enumerate(sectores, start=1):
    ws.append([i, sector, 1])  # linea_id = 1 para todos

# Guardar el archivo
wb.save("sectores1.xlsx")
print("Archivo 'sectores.xlsx' generado exitosamente.")